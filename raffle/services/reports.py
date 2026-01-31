"""Excel report builders for administrative exports."""

from __future__ import annotations

from io import BytesIO

from django.db.models import Count
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..rooms import RoomDirectory
from .summary import get_coupon_room_summary


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TABLE_BORDER = Border(
    left=Side(border_style="thin", color="D9D9D9"),
    right=Side(border_style="thin", color="D9D9D9"),
    top=Side(border_style="thin", color="D9D9D9"),
    bottom=Side(border_style="thin", color="D9D9D9"),
)
ALT_ROW_FILL = PatternFill(start_color="F2F6FB", end_color="F2F6FB", fill_type="solid")


def build_coupon_report_workbook(coupons) -> Workbook:
    """Build a styled workbook with coupon details and room summary."""

    workbook = Workbook()
    _build_summary_sheet(workbook, coupons)
    _build_detail_sheet(workbook, coupons)
    return workbook


def build_room_report_workbook(coupons) -> Workbook:
    """Build a workbook grouped by room with totals and detailed rows."""

    workbook = Workbook()
    _build_room_summary_sheet(workbook, coupons)
    _build_room_detail_sheet(workbook, coupons)
    return workbook


def build_daily_report_workbook(coupons) -> Workbook:
    """Build a workbook grouped by day with totals and detailed rows."""

    workbook = Workbook()
    _build_daily_summary_sheet(workbook, coupons)
    _build_daily_detail_sheet(workbook, coupons)
    return workbook


def render_workbook_response(workbook: Workbook, filename: str) -> HttpResponse:
    """Return an HTTP response streaming the workbook as an XLSX file."""

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _build_summary_sheet(workbook: Workbook, coupons) -> None:
    """Create the summary tab grouped by room."""

    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append(["Sala", "Cupones", "Última emisión"])
    _style_header_row(sheet)

    summary = get_coupon_room_summary(coupons)
    for index, item in enumerate(summary, start=2):
        last_coupon = coupons.filter(room_id=item["room"].id).order_by("-scanned_at").first()
        last_date = _format_datetime(last_coupon.scanned_at) if last_coupon else "-"
        sheet.append([item["room"].name, item["total"], last_date])
        _style_data_row(sheet, index)

    _autosize_columns(sheet)
    sheet.freeze_panes = "A2"


def _build_detail_sheet(workbook: Workbook, coupons) -> None:
    """Create the detailed tab with each coupon row."""

    sheet = workbook.create_sheet("Cupones")
    headers = [
        "Código",
        "Participante",
        "DNI",
        "Teléfono",
        "Correo",
        "Sala",
        "Terminal",
        "Origen",
        "Fecha escaneo",
    ]
    sheet.append(headers)
    _style_header_row(sheet)

    for index, coupon in enumerate(coupons, start=2):
        sheet.append(
            [
                coupon.code,
                f"{coupon.person.first_name} {coupon.person.last_name}".strip(),
                coupon.person.id_number,
                coupon.person.phone,
                coupon.person.email,
                RoomDirectory.get(coupon.room_id).name,
                coupon.terminal_name,
                coupon.get_source_display(),
                _format_datetime(coupon.scanned_at),
            ]
        )
        _style_data_row(sheet, index)

    _autosize_columns(sheet)
    sheet.freeze_panes = "A2"


def _style_header_row(sheet) -> None:
    """Apply brand styling to the header row."""

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = TABLE_BORDER


def _style_data_row(sheet, row_index: int) -> None:
    """Add subtle styling to data rows for readability."""

    for cell in sheet[row_index]:
        cell.border = TABLE_BORDER
        cell.alignment = Alignment(vertical="center")
        if row_index % 2 == 0:
            cell.fill = ALT_ROW_FILL


def _autosize_columns(sheet) -> None:
    """Resize columns based on their longest value."""

    for column_cells in sheet.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column].width = max(14, min(max_length + 2, 50))


def _format_datetime(value) -> str:
    """Format datetimes with local time awareness."""

    if not value:
        return ""
    if timezone.is_naive(value):
        localized = timezone.make_aware(value, timezone.get_current_timezone())
    else:
        localized = timezone.localtime(value)
    return localized.strftime("%d/%m/%Y %H:%M")


def _format_date(value) -> str:
    """Format dates using the current timezone."""

    if not value:
        return ""
    timezone_ref = timezone.get_current_timezone()
    localized = (
        timezone.localtime(value, timezone_ref)
        if timezone.is_aware(value)
        else timezone.make_aware(value, timezone_ref)
    )
    return localized.date().strftime("%d/%m/%Y")


def _build_room_summary_sheet(workbook: Workbook, coupons) -> None:
    """Create the room summary tab with counts and last scanned date."""

    sheet = workbook.active
    sheet.title = "Por sala"
    sheet.append(["Sala", "Cupones", "Última emisión"])
    _style_header_row(sheet)

    summary = get_coupon_room_summary(coupons)
    for index, item in enumerate(summary, start=2):
        last_coupon = coupons.filter(room_id=item["room"].id).order_by("-scanned_at").first()
        last_date = _format_datetime(last_coupon.scanned_at) if last_coupon else "-"
        sheet.append([item["room"].name, item["total"], last_date])
        _style_data_row(sheet, index)

    _autosize_columns(sheet)
    sheet.freeze_panes = "A2"


def _build_room_detail_sheet(workbook: Workbook, coupons) -> None:
    """Create the detail tab sorted by room."""

    sheet = workbook.create_sheet("Cupones por sala")
    headers = [
        "Sala",
        "Código",
        "Participante",
        "DNI",
        "Teléfono",
        "Correo",
        "Terminal",
        "Origen",
        "Fecha escaneo",
    ]
    sheet.append(headers)
    _style_header_row(sheet)

    for index, coupon in enumerate(coupons.order_by("room_id", "-scanned_at"), start=2):
        sheet.append(
            [
                RoomDirectory.get(coupon.room_id).name,
                coupon.code,
                f"{coupon.person.first_name} {coupon.person.last_name}".strip(),
                coupon.person.id_number,
                coupon.person.phone,
                coupon.person.email,
                coupon.terminal_name,
                coupon.get_source_display(),
                _format_datetime(coupon.scanned_at),
            ]
        )
        _style_data_row(sheet, index)

    _autosize_columns(sheet)
    sheet.freeze_panes = "A2"


def _build_daily_summary_sheet(workbook: Workbook, coupons) -> None:
    """Create the daily summary tab with totals per day."""

    sheet = workbook.active
    sheet.title = "Por día"
    sheet.append(["Fecha", "Cupones"])
    _style_header_row(sheet)

    daily_totals = (
        coupons.annotate(day=TruncDate("scanned_at"))
        .values("day")
        .annotate(total=Count("id"))
        .order_by("day")
    )
    for index, item in enumerate(daily_totals, start=2):
        sheet.append([item["day"].strftime("%d/%m/%Y") if item["day"] else "-", item["total"]])
        _style_data_row(sheet, index)

    _autosize_columns(sheet)
    sheet.freeze_panes = "A2"


def _build_daily_detail_sheet(workbook: Workbook, coupons) -> None:
    """Create the detail tab sorted by day."""

    sheet = workbook.create_sheet("Cupones por día")
    headers = [
        "Fecha",
        "Código",
        "Participante",
        "DNI",
        "Teléfono",
        "Correo",
        "Sala",
        "Terminal",
        "Origen",
        "Hora",
    ]
    sheet.append(headers)
    _style_header_row(sheet)

    for index, coupon in enumerate(coupons.order_by("scanned_at"), start=2):
        sheet.append(
            [
                _format_date(coupon.scanned_at),
                coupon.code,
                f"{coupon.person.first_name} {coupon.person.last_name}".strip(),
                coupon.person.id_number,
                coupon.person.phone,
                coupon.person.email,
                RoomDirectory.get(coupon.room_id).name,
                coupon.terminal_name,
                coupon.get_source_display(),
                _format_datetime(coupon.scanned_at),
            ]
        )
        _style_data_row(sheet, index)

    _autosize_columns(sheet)
    sheet.freeze_panes = "A2"
