from datetime import date
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from ..models import Coupon, Person, VoucherScan
from ..services.voucher_validation import VoucherValidationResult
from utils.printers import build_escpos_bytes


class VoucherScanModelTests(TestCase):
    def test_room_name_property_returns_human_readable_name(self):
        """Ensure the voucher scan exposes the readable room name."""

        person = Person.objects.create(
            first_name="Test",
            last_name="User",
            id_number="12345678",
            phone="5551234",
            birth_date=date(1990, 1, 1),
        )
        scan = VoucherScan.objects.create(
            code="VOUCHER-001",
            person=person,
            room_id=1,
            source=Coupon.ENTRY,
        )

        self.assertEqual(scan.room_name, "SCN")


class EntryViewTests(TestCase):
    def setUp(self):
        self.person = Person.objects.create(
            first_name="Sample",
            last_name="User",
            id_number="87654321",
            phone="5559876",
            birth_date=date(1990, 5, 17),
        )
        self.validation_patcher = patch("raffle.controllers.public.validate_voucher_code")
        self.mock_validate = self.validation_patcher.start()
        self.mock_validate.return_value = VoucherValidationResult(True, "OK")

    def tearDown(self):
        self.validation_patcher.stop()

    def test_entry_scan_generates_coupon_and_triggers_print(self):
        """Ensure scanning a voucher generates a coupon and prints it."""

        payload = {
            "id_number": self.person.id_number,
            "room_id": "1",
            "voucher_code": "scan-001",
        }
        with patch("raffle.controllers.public.print_coupon_backend") as mock_print:
            response = self.client.post(reverse("raffle:entry"), data=payload)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], reverse("raffle:home"))

        scan_exists = VoucherScan.objects.filter(code="SCAN-001", person=self.person).exists()
        self.assertTrue(scan_exists)

        coupons = Coupon.objects.filter(person=self.person, source=Coupon.ENTRY)
        self.assertEqual(coupons.count(), 1)
        mock_print.assert_called_once()
        printed_coupon = mock_print.call_args[0][0]
        self.assertEqual(printed_coupon.id, coupons.first().id)

    def test_entry_rejects_scans_within_two_hours_for_same_person(self):
        """Block scans if the participant used a voucher within two hours."""

        payload = {
            "id_number": self.person.id_number,
            "room_id": "1",
            "voucher_code": "scan-001",
        }
        with patch("raffle.controllers.public.print_coupon_backend"):
            self.client.post(reverse("raffle:entry"), data=payload)

        follow_up_payload = {
            "id_number": self.person.id_number,
            "room_id": "1",
            "voucher_code": "scan-002",
        }
        response = self.client.post(reverse("raffle:entry"), data=follow_up_payload)

        scans = VoucherScan.objects.filter(person=self.person)
        self.assertEqual(scans.count(), 1)
        self.assertContains(
            response,
            "Solo puedes escanear un ticket cada dos horas.",
            status_code=200,
        )

    def test_entry_enforces_daily_coupon_limit(self):
        """Prevent issuing more than 10 entry coupons per day."""

        for index in range(10):
            Coupon.objects.create(
                person=self.person,
                code=f"LIMIT-{index}",
                source=Coupon.ENTRY,
                room_id=1,
                scanned_at=timezone.now(),
            )

        payload = {
            "id_number": self.person.id_number,
            "room_id": "1",
            "voucher_code": "limit-011",
        }
        response = self.client.post(reverse("raffle:entry"), data=payload)

        self.assertEqual(
            Coupon.objects.filter(person=self.person, source=Coupon.ENTRY).count(),
            10,
        )
        self.assertContains(
            response,
            "Se alcanzó el límite diario de 10 cupones para esta persona.",
            status_code=200,
        )


class CouponPrintingTests(TestCase):
    def test_coupon_payload_contains_trailing_blank_lines(self):
        """Ensure coupon printing includes trailing blank lines before cutting."""

        person = Person.objects.create(
            first_name="Print",
            last_name="Tester",
            id_number="11112222",
            phone="5554444",
            birth_date=date(1985, 7, 1),
        )
        coupon = Coupon.objects.create(
            person=person,
            code="PRINT-001",
            source=Coupon.ENTRY,
            room_id=1,
        )

        payload = build_escpos_bytes(coupon)
        self.assertTrue(payload.endswith(b"\n\n\n\n\n\x1d\x56\x00"))


class CouponExportTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.admin = User.objects.create_user(
            username="admin_export", password="secret", role=User.Role.ADMIN
        )
        self.client.force_login(self.admin)

        self.person = Person.objects.create(
            first_name="Export",
            last_name="User",
            id_number="22223333",
            phone="5550000",
            birth_date=date(1980, 3, 3),
        )
        Coupon.objects.create(
            person=self.person,
            code="EXPORT-001",
            source=Coupon.ENTRY,
            room_id=1,
            terminal_name="Terminal-1",
            scanned_at=timezone.now(),
        )

    def test_admin_can_export_coupons_to_excel(self):
        """Ensure the export endpoint returns a styled Excel workbook."""

        response = self.client.get(reverse("raffle_admin:coupons_export"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = load_workbook(BytesIO(response.content))
        self.assertIn("Cupones", workbook.sheetnames)
        detail_sheet = workbook["Cupones"]
        self.assertGreaterEqual(detail_sheet.max_row, 2)
